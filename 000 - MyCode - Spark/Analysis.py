""""
Program name: Analysis.py
----------------------------------------------------------------------
Program elaborated to fetch data from a MySQL database and analyze it, numerically and graphically.
----------------------------------------------------------------------
Description:
Python program to analyze waveform data stored in a MySQL database.
It fetches waveform data, analyzes it for various metrics (mean, median, number of peaks, amplitude, period), and saves the results in an Excel file.
The program also generates various plots of the waveforms, including raw, vertically aligned, and filtered by amplitude.
It includes functions to plot the mean waveform and aligned waveforms based on peaks.
The aim of this program is to highlight patterns and trends in the waveform data, providing insights into the behavior of the system being monitored.
----------------------------------------------------------------------
Structure:

| ml_base_data.py - Machine learning base data file (Fetches data from the database and allows to categorize it, providing feedback for machine learning model input).
| ml_model.py - Machine learning model file (Based on the feedback trains a machine learning model to predict spark events).
|
| main.py - Main program file
|__ Acquisition.py - Acquires data from an oscilloscope and saves it in a MySQL database.
|__ Evaluation.py - Evaluate database data file (Fetches data from the database and evaluates it using the trained machine learning model).
|__ Analysis.py - Analyzes the spark data and generates a report with Graphical and Analytical results.
|_____ MetricAnalysis.py - Provides an analysis on the impact of mechanical defaults on the spark events, based on test metrics (validation).

Environment:
  1. Python 3.12.7
  2. mysql-connector-python 2.2.9
  3. matplotlib 3.4.3
  4. numpy 1.21.0
  5. csv 1.0.0
  6. os 0.1.0
  7. cv2 4.5.3
  8. re 4.3.0
  9. scipy 1.7.1
  10. ast 2.0.0
  11. openpyxl 3.0.9
  12. warnings 4.0.3
  13. seaborn 0.11.2
  14. time 4.0.0

Version: 1.0
Modified on Mar 31th 2025
Author: Maria Rodrigues
"""

import mysql.connector
import numpy as np
from scipy.signal import find_peaks
import re
import ast
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
import warnings
import matplotlib
matplotlib.use('TkAgg')  # Ensure it's interactive
import matplotlib.pyplot as plt
import seaborn as sns
import os
import time
from Acquisition import periods as periods

warnings.filterwarnings("ignore", category=RuntimeWarning)

########################## Defining Plot Settings ########################## 
 
# Set the color palette
blue_palette = sns.color_palette("Blues", 8)
red_palette = sns.color_palette("Reds", 8)
green_palette = sns.color_palette("Greens", 8)
grey_palette = sns.color_palette("Greys", 8)
 
custom_palette = [blue_palette[6], blue_palette[2], red_palette[5], red_palette[2], green_palette[5], green_palette[2], grey_palette[7], grey_palette[2]]
 
# Set the style
sns.set_style("ticks")
 
# Set default figure size
plt.rcParams['figure.figsize'] = (12, 8)
 
# Set text sizes
plt.rcParams['axes.titlesize'] = 36
plt.rcParams['axes.labelsize'] = 34
plt.rcParams['xtick.labelsize'] = 30
plt.rcParams['ytick.labelsize'] = 30
plt.rcParams['legend.fontsize'] = 30
plt.rcParams['font.size'] = 30
plt.rcParams['axes.linewidth'] = 1.0
 
# Use LaTeX for rendering text
plt.rcParams['text.usetex'] = True
plt.rcParams['font.family'] = 'serif'
plt.rcParams['font.serif'] = 'CMU'
plt.rcParams['text.latex.preamble'] = r'\usepackage{amsmath}'

save_path = r"C:\Users\Maria Rodrigues\Desktop\Investigação ILLIANCE BOSCH\003 - Lab Tests\002 - Spark Test 19.03.25\PDFs"

########################## Database connection ##########################

# Function to fetch waveform data from the database
def fetch_waveform_data(period_start, period_end):
    db = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="sparktest"
    )
    cursor = db.cursor()
    
    # Fetch waveform data for the specified period for all IDs that are sparks
    query = """ 
    SELECT id, waveform_data_ch1 FROM oscilloscope_data WHERE id IN (SELECT id FROM sparks) AND id BETWEEN %s AND %s;
    """
    cursor.execute(query, (period_start, period_end))
    
    wave_datalist = []
    id_list = []
    
    for row in cursor.fetchall():
        waveform_str = row[1].strip()
        waveform_str = re.sub(r"([0-9e+\-.]+)(?=\s+([0-9e+\-.]))", r"\1, ", waveform_str)
        
        try:
            waveform = ast.literal_eval(waveform_str)
            if isinstance(waveform, list):
                wave_datalist.append(waveform)
                id_list.append(row[0])
        except Exception as e:
            print(f"Error parsing waveform for ID {row[0]}: {e}")
    
    cursor.close()
    db.close()
    return wave_datalist, id_list

########################## For numerical analysis ##########################
# Function to analyze waveforms (with threshold)
def analyze_waveforms_with_threshold(wave_datalist, id_list, amplitude_threshold=0.35):
    results = []
    valid_averages, valid_medians, valid_num_peaks, valid_ampl, valid_periods = [], [], [], [], []

    for waveform, id_value in zip(wave_datalist, id_list):
        peaks, _ = find_peaks(waveform, height=amplitude_threshold)
        num_peaks_value = len(peaks)
        amplitude_value = np.max(waveform) - np.min(waveform)
        
        # Check for valid peaks before calculating period
        if num_peaks_value > 1:
            period_value = peaks[-1] - peaks[0]
        else:
            period_value = np.nan  # or you could use 'Null'
        
        # Ensure we only process waveforms with valid peaks
        if num_peaks_value >0:
            try:
                average_value = np.mean(waveform)
            except:
                average_value = "error - mean of empty slice"

            try:
                median_value = np.median(waveform)
            except:
                median_value = "error - mean of empty slice"
            
            if amplitude_value >= 1:
                valid_averages.append(average_value)
                valid_medians.append(median_value)
                valid_num_peaks.append(num_peaks_value)
                valid_ampl.append(amplitude_value)
                if num_peaks_value != 0:
                    valid_periods.append(period_value)

            results.append({
                'ID': id_value, 'Average': average_value, 'Median': median_value,
                'Number of Peaks': num_peaks_value if amplitude_value >= 1 else 'Null',
                'Amplitude': amplitude_value if amplitude_value >= 1 else 'Null',
                'Period': period_value if (amplitude_value >= 1 and num_peaks_value != 0) else 'Null',
            })
        else:
            results.append({
                'ID': id_value, 'Average': 'Null', 'Median': 'Null',
                'Number of Peaks': 'Null', 'Amplitude': 'Null', 'Period': 'Null',
            })

    return results, valid_averages, valid_medians, valid_num_peaks, valid_ampl, valid_periods

# Function for analysis without threshold
def analyze_waveforms_raw_threshold(wave_datalist, id_list):
    results, averages, medians, num_peaks, ampl, periods = [], [], [], [], [], []

    for waveform, id_value in zip(wave_datalist, id_list):
        peaks, _ = find_peaks(waveform)
        num_peaks_value = len(peaks)
        amplitude_value = np.max(waveform) - np.min(waveform)
        
        # Check for valid peaks before calculating period
        if num_peaks_value> 1:
            period_value = peaks[-1] - peaks[0]
        else:
            period_value = np.nan  # or 'Null'

        # Ensure only process waveforms with valid peaks
        if num_peaks_value >0:
            try:
                average_value = np.mean(waveform)
            except:
                average_value = "error - mean of empty slice"

            try:
                median_value = np.median(waveform)
            except:
                median_value = "error - mean of empty slice"
        
            results.append({
                'ID': id_value, 'Average': average_value, 'Median': median_value,
                'Number of Peaks': num_peaks_value, 'Amplitude': amplitude_value, 'Period': period_value,
            })
            
            averages.append(average_value)
            medians.append(median_value)
            num_peaks.append(num_peaks_value)
            ampl.append(amplitude_value)
            periods.append(period_value)
        else:
            results.append({
                'ID': id_value, 'Average': 'Null', 'Median': 'Null',
                'Number of Peaks': 'Null', 'Amplitude': 'Null', 'Period': 'Null',
            })

    return results, averages, medians, num_peaks, ampl, periods

# Save results to Excel
def save_to_excel(results_with_threshold, results_raw, period_start, period_end, avg_with_threshold, avg_raw_threshold, wave_datalist, id_list):
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "1 - Waveform Strings"
    ws1.append(['ID', 'Waveform String'])

    for id_value, waveform in zip(id_list, wave_datalist):
        ws1.append([id_value, str(waveform)])

    ws2 = wb.create_sheet(title="2 - Without Threshold")
    headers = ['ID', 'Average', 'Median', 'Number of Peaks', 'Amplitude', 'Period']
    ws2.append(headers)
    
    for result in results_raw:
        ws2.append([result['ID'], result['Average'], result['Median'], result['Number of Peaks'], result['Amplitude'], result['Period']])

    ws3 = wb.create_sheet(title="3 - Waveform Analysis")
    ws3.append(headers)

    for result in results_with_threshold:
        ws3.append([result['ID'], result['Average'], result['Median'], result['Number of Peaks'], result['Amplitude'], result['Period']])

    ws4 = wb.create_sheet(title="4 - Summary")    
    ws4.append(["Metric", "With Threshold", "Without Threshold"])
    metric_names = ["Mean of Averages", "Mean of Medians", "Mean of Number of Peaks", "Mean of ampl", "Mean of Periods"]

    for i, metric in enumerate(metric_names):
        ws4.append([metric, avg_with_threshold[i] * (1e-3 if "Period" in metric else 1), avg_raw_threshold[i] * (1e-3 if "Period" in metric else 1)])

    # Starting position for the first chart
    row_offset = 2  

    # Loop through each row (2 to 7) and create one chart per metric
    for i in range(2, 8):  
        chart = LineChart()
        chart.title = ws4[f"A{i}"].value  # Use metric name as title
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Condition"  # Label for X-axis

        # Ensure axes values (ticks) are visible
        chart.x_axis.majorTickMark = "out"  # Show X-axis tick marks
        chart.y_axis.majorTickMark = "out"  # Show Y-axis tick marks

        # Remove legend
        chart.legend = None  # This disables the legend

        # X-values: Always B1 → C1 (With Threshold, Without Threshold)
        categories = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=1)
        
        # Y-values: Only row i, columns B and C
        values = Reference(ws4, min_col=2, max_col=3, min_row=i, max_row=i)
        
        # Create a single series for this row
        series = Series(values, title=ws4[f"A{i}"].value)
        chart.append(series)  # Add to chart
        chart.set_categories(categories)  # Set X-axis values (B1 → C1)

        # Position chart in column E, leaving 15 rows between each
        ws4.add_chart(chart, f"E{row_offset}")
        row_offset += 15  # Space between charts
        
    # Save workbook
    output_filename = f"waveform_analysis_{period_start}_{period_end}.xlsx"
    wb.save(output_filename)
    print(f"Results saved to {output_filename}")

########################## For graphical analysis ##########################

# Utility function to center waveform
def center_waveform(waveform):
    return np.array(waveform) - np.mean(waveform)

# Utility function to filter waveforms by amplitude
def filter_by_amplitude(waveform, threshold):
    amplitude = np.max(waveform) - np.min(waveform)
    return amplitude > threshold

# Plot all waveforms exactly as they are
def plot_all_raw(wave_datalist, ids, period_start, period_end):
    plt.figure(figsize=(12, 8))
    for waveform, id_value in zip(wave_datalist, ids):
        plt.plot(waveform, alpha=0.3, color = custom_palette[0])
    
   # plt.title("All Waveforms Overlapped (Raw)")
    plt.xlabel("Sample Index")
    plt.ylabel("Amplitude (V)")
    plt.grid(True)
    plt.tight_layout()
    
    # Save with IDs in filename
    plt.savefig(f"raw_waveforms_{period_start}_to_{period_end}.png")
    #plt.savefig(os.path.join(save_path, f'raw_waveforms_{period_start}_to_{period_end}.pdf'), dpi = 300)

    plt.close()
    #print(f"Graph saved as 'raw_waveforms_{period_start}_to_{period_end}.png'.")

# Plot all waveforms overlapped, vertically aligned (start at 0)
def plot_vertically_aligned(wave_datalist, ids, period_start, period_end):
    plt.figure(figsize=(12, 8))
    for waveform, id_value in zip(wave_datalist, ids):
        aligned = np.array(waveform) - waveform[0]  # Align start to 0
        plt.plot(aligned, alpha=0.3, color = custom_palette[0])
    
   # plt.title("All Waveforms Vertically Aligned (Start at 0)")
    plt.xlabel("Sample Index")
    plt.ylabel("Amplitude (V)")
    plt.grid(True)
    plt.tight_layout()
    
    # Save with IDs in filename
    plt.savefig(f"vertically_aligned_waveforms_{period_start}_to_{period_end}.png")
    #plt.savefig(os.path.join(save_path, f'vertically_aligned_waveforms_{period_start}_to_{period_end}.pdf'), dpi = 300)
    plt.close()
    #print(f"Graph saved as 'vertically_aligned_waveforms_{period_start}_to_{period_end}.png'.")


# Predefined legend labels for 8 tests
legend_labels = ["Ref.", "T.1", "T.2", "T.3", "T.4", "T.5", "T.6", "T.7"]
all_mean_waveforms = []  # Store tuples of (label, mean_waveform)

def plot_filtered_and_mean(wave_datalist, ids, period_start, period_end):
    # Center and filter waveforms by amplitude > 0.5
    valid_waveforms = [center_waveform(waveform) for waveform in wave_datalist if filter_by_amplitude(waveform, 0.5)]

    if not valid_waveforms:
        print("No waveforms with amplitude greater than 0.5 found.")
        return

    plt.figure(figsize=(12, 8))

    # Plot each waveform
    for wf in valid_waveforms:
        plt.plot(wf, alpha=0.3, color=custom_palette[0])

    # Plot and store the mean waveform
    mean_waveform = np.mean(valid_waveforms, axis=0)
    # Use number of collected means so far to assign label
    label_index = len(all_mean_waveforms)
    label = legend_labels[label_index] if label_index < len(legend_labels) else f"extra_{label_index}"
    all_mean_waveforms.append((label, mean_waveform))

    plt.plot(mean_waveform, color='red', linewidth=3, label='Mean Waveform')

    plt.xlabel("Sample Index")
    plt.ylabel("Amplitude (V)")
    plt.legend(loc='upper right')
    plt.grid(True)
    plt.tight_layout()

    # Save individual graph
    plt.savefig(f"filtered_mean_centered_waveforms_{period_start}_to_{period_end}.png")
    #plt.savefig(os.path.join(save_path, f'filtered_mean_centered_waveforms_{period_start}_to_{period_end}.pdf'), dpi=300)

    plt.close()
    #print(f"Graph saved as 'filtered_mean_centered_waveforms_{period_start}_to_{period_end}.png'.")
    
def plot_all_means_overlay():
    if not all_mean_waveforms:
        print("No mean waveforms to overlay.")
        return

    plt.figure(figsize=(12, 8))

    # Light Red and Light Blue colors, from dark to light
    colors = [
        '#1f77b4', '#d62728',  # Dark blue, Dark red
        '#4292c6', '#e84a5f',  # Medium blue, Medium red
        '#6baed6', '#f27f79',  # Lighter blue, Lighter red
        '#9ecae1', '#fbb6b6',  # Light blue, Very light red
        '#c6dbef', '#fdd8d8',  # Pale blue, Pale red
        '#e6f0f6', '#ffe5e5',  # Very pale blue, Very pale red
        '#d0e4f1', '#ffcccc',  # Soft blue, Soft red
        '#bdd7ea'             # Lightest blue
    ]

    handles = []
    labels = []

    for i, (label, mean_wf) in enumerate(all_mean_waveforms):
        # Pick alternating colors from the color list
        color = colors[i % len(colors)]
        # Plot the waveform, no y-offset to overlay them
        line, = plt.plot(mean_wf, label=label, color=color, linewidth=2)
        handles.append(line)
        labels.append(label)

    plt.xlabel("Sample Index")
    plt.ylabel("Amplitude (V)")

    # Move the legend above the graph
    plt.legend(handles, labels, loc='upper center', bbox_to_anchor=(0.5, 1.15), ncol=4, frameon=False)

    # Adjust layout to make room for the legend
    plt.subplots_adjust(top=1.1)
    plt.grid(True)

    # Save the graph as PNG and PDF
    plt.savefig("overlay_all_mean_waveforms.png", bbox_inches='tight')
    #plt.savefig(os.path.join(save_path, "overlay_all_mean_waveforms.pdf"), dpi=300, bbox_inches='tight')
    plt.close()
    #print("Overlay of all mean waveforms saved.")
    
# Plot only waveforms with amplitude Bigger Than 1, centered, aligned at 0 with mean in red
def plot_filtered_amp1(wave_datalist, ids, period_start, period_end):
    # Center and filter waveforms by amplitude Bigger Than 1
    valid_waveforms = [center_waveform(waveform) for waveform in wave_datalist if filter_by_amplitude(waveform,1)]
    
    if not valid_waveforms:
        print("No waveforms with amplitude greater than 1 found.")
        return
    
    plt.figure(figsize=(12, 8))
    
    # Plot each waveform
    for wf in valid_waveforms:
        plt.plot(wf, alpha=0.3, color = custom_palette[0])
    
    # Plot the mean waveform
    mean_waveform = np.mean(valid_waveforms, axis=0)
    plt.plot(mean_waveform, color='red', linewidth=3, label='Mean Waveform')
    
   # plt.title("Centered Waveforms with Amplitude Bigger Than 2 and Mean in Red")
    plt.xlabel("Sample Index")
    plt.ylabel("Amplitude (V)")
    plt.legend(loc='upper right')
    plt.grid(True)
    plt.tight_layout()
    
    # Save with IDs in filename
    plt.savefig(f"filtered_amp_bigger_than1_{period_start}_to_{period_end}.png")
    #plt.savefig(os.path.join(save_path, f'filtered_amp_bigger_than1_{period_start}_to_{period_end}.pdf'), dpi = 300)

    plt.close()
    #print(f"Graph saved as 'filtered_amp_bigger_than1_{period_start}_to_{period_end}.png'.")
    
    
# Plot only waveforms with amplitude Bigger Than 2, centered, aligned at 0 with mean in red
def plot_filtered_amp2(wave_datalist, ids, period_start, period_end):
    # Center and filter waveforms by amplitude Bigger Than 2
    valid_waveforms = [center_waveform(waveform) for waveform in wave_datalist if filter_by_amplitude(waveform,2)]
    
    if not valid_waveforms:
        print("No waveforms with amplitude greater than 2 found.")
        return
    
    plt.figure(figsize=(12, 8))
    
    # Plot each waveform
    for wf in valid_waveforms:
        plt.plot(wf, alpha=0.3, color = custom_palette[0])
    
    # Plot the mean waveform
    mean_waveform = np.mean(valid_waveforms, axis=0)
    plt.plot(mean_waveform, color='red', linewidth=3, label='Mean Waveform')
    
   # plt.title("Centered Waveforms with Amplitude Bigger Than 2 and Mean in Red")
    plt.xlabel("Sample Index")
    plt.ylabel("Amplitude (V)")
    plt.legend(loc='upper right')
    plt.grid(True)
    plt.tight_layout()
    
    # Save with IDs in filename
    plt.savefig(f"filtered_amp_bigger_than2_{period_start}_to_{period_end}.png")
    #plt.savefig(os.path.join(save_path, f'filtered_amp_bigger_than2_{period_start}_to_{period_end}.pdf'), dpi = 300)

    plt.close()
    #print(f"Graph saved as 'filtered_amp_bigger_than2_{period_start}_to_{period_end}.png'.")


# Plot all waveforms aligned by their highest point
def plot_aligned_by_peaks(wave_datalist, ids, period_start, period_end):
    aligned_waveforms = []
    
    for waveform in wave_datalist:
        aligned = np.array(waveform) - waveform[0]  # Align start to 0
        peak_index = np.argmax(aligned)
        x_aligned = np.arange(len(aligned)) - peak_index
        aligned_waveforms.append((x_aligned, aligned))
    
    plt.figure(figsize=(12, 8))
    
    for x_aligned, wf in aligned_waveforms:
        plt.plot(x_aligned, wf, color='black', alpha=0.5)
    
    # Interpolate the waveforms to common x-axis and calculate mean
    common_x = np.arange(len(wave_datalist[0])) - len(wave_datalist[0]) // 2
    aligned_interpolated = [np.interp(common_x, x_aligned, wf) for x_aligned, wf in aligned_waveforms]
    
    mean_waveform = np.mean(aligned_interpolated, axis=0)
    plt.plot(common_x, mean_waveform, color='red', linewidth=3, label='Mean Waveform')
    
   # plt.title("Waveforms Aligned by Peaks with Mean in Red")
    plt.xlabel("Sample Index (Peak at 0)")
    plt.ylabel("Amplitude (V)")
    plt.legend(loc='upper right')
    plt.grid(True)
    plt.tight_layout()
    
    # Save with IDs in filename
    plt.savefig(f"aligned_by_peaks_waveforms_{period_start}_to_{period_end}.png")
    #plt.savefig(os.path.join(save_path, f'aligned_by_peaks_waveforms_{period_start}_to_{period_end}.pdf'), dpi = 300)

    plt.close()
    #print(f"Graph saved as 'aligned_by_peaks_waveforms_{period_start}_to_{period_end}.png'.")


# Main function
def main():

    for period_start, period_end in periods:
        
        wave_datalist, id_list = fetch_waveform_data(period_start, period_end)
        
        results_raw , avgs_raw, medians_raw, num_peaks_raw, ampl_raw, periods_raw = analyze_waveforms_raw_threshold(wave_datalist, id_list)
        results_with_threshold, avgs_with, medians_with, num_peaks_with, ampl_with, periods_with = analyze_waveforms_with_threshold(wave_datalist, id_list)
        
        avg_raw_threshold = [ np.mean(avgs_raw), np.mean(medians_raw), np.mean(num_peaks_raw), np.mean(ampl_raw), np.mean(periods_raw)]
        avg_with_threshold = [ np.mean(avgs_with), np.mean(medians_with), np.mean(num_peaks_with), np.mean(ampl_with), np.mean(periods_with)]

        save_to_excel(results_with_threshold, results_raw, period_start, period_end, avg_with_threshold, avg_raw_threshold, wave_datalist, id_list)
        
        wave_datalist, id_list = fetch_waveform_data(period_start, period_end)
        
        plot_all_raw(wave_datalist, id_list, period_start, period_end)  # All raw waveforms
        plot_vertically_aligned(wave_datalist, id_list, period_start, period_end)  # All aligned vertically (start = 0)
        plot_aligned_by_peaks(wave_datalist, id_list, period_start, period_end)  # All aligned by peaks
        plot_filtered_and_mean(wave_datalist, id_list, period_start, period_end)  # Filtered by amplitude Bigger Than 0.5 with mean
        plot_filtered_amp1(wave_datalist, id_list, period_start, period_end)  # Filtered by amplitude Bigger Than 1 with mean
        plot_filtered_amp2(wave_datalist, id_list, period_start, period_end)  # Filtered by amplitude Bigger Than 2 with mean
    
    for period_start, period_end in periods:    
        plot_all_means_overlay()
